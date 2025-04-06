from pdf2image import convert_from_path
import pytesseract
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

def convert_freedom_pdf_to_excel(pdf_path, excel_path):
    images = convert_from_path(pdf_path)
    wb = openpyxl.Workbook()
    ws = wb.active

    row_idx = 1
    for img in images:
        text = pytesseract.image_to_string(img, lang="rus+eng")
        lines = text.splitlines()
        for line in lines:
            if line.strip():
                ws.cell(row=row_idx, column=1, value=line.strip())
                row_idx += 1

    wb.save(excel_path)

def save_transactions_to_excel(transactions: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Дата", "Описание", "Сумма", "Тип", "Источник"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for tx in transactions:
        ws.append([
            tx.get("date", ""),
            tx.get("description", ""),
            float(tx.get("amount", 0)),  # 💡 явно преобразуем в число
            tx.get("type", ""),
            tx.get("source", "")
        ])

    wb.save(output_path)